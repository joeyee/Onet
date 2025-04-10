'''
In default zy3_trainset, many images contains no clouds and fused with ice/snow.
Here, we filter the trainset to select the images with correct clouds.
and use these segmentated clouds to synthesize the clouds on the terrains wihtout clouds,
especially the terrains only with ice/snow.

Created by ZhouYi@Linghai_Dalian on 2024/03/07
'''
import os
import numpy as np
import matplotlib.pyplot as plt
import gc
from datetime import datetime
import torch
import torch.nn as nn
import utils_20231218 as uti
import configs.config_tip2022_20230411           as conf_model
import dataloader.zy3_cloud_thumbnailv5_20240304 as cloud_model_zy3
import uti_zy3_test_20240123 as uti_zy3_test
#import Onet_vanilla_outc_20240305              as onet_vanilla_outc
import torch.nn.functional as F
import copy
import logging
import openpyxl
from PIL import Image

from tqdm import tqdm
torch.manual_seed(1981)
np.random.seed(1981)
torch.set_default_dtype(torch.float32)

bright_terrain_only_tid_list = ['1712110579', '1712119384', '17012277836044602',
                                '14120165633954060', '16102400785553324',
                                '17040744346354862', '17042537416415819',
                                '14112183753944996', '14112058503919781',
                                '17032328336275058', '17041101236382351',
                                '17042778266418650', '17032348016275067',
                                '17020878486096639', '1706157731',
                                '17101849847322975', '14112058503919781',
                                '17091422397172296',  # not find 1709142239712296 in trainset
                                '15012121584084790', '14121797223988205', '17120177657504539',
                                '17042552596415577', '1710177084',
                                '14120165633954060', '1210290160586232',  # ice/snow land
                                '1711213249', '1712019369', '1712011771', '1712115356',
                                '1712115356', '1712118502', '1609232855', '1609230301', '1712118687', '1712010073',
                                '1711218658', '1712125418', '1710174747', '1609238908',  # sand/rock land
                                '1712125418', '1712112118'  # suburb
                                ]
thin_clouds_only_tid_list = ['1711210256', '1711211564', '1711212921', '1711210256',
                             '1702105821', '1710171813', '17062222776751076', '1712043142',
                             '1710172901', '1706156981', '1706159113', '17062206586751086',
                             '1706150953', '1702105574', '1712075643', '1712077273',
                             '1711215376', '17042089736367046', '17041105826425179', '1706158902',
                             ]


def seg_on_zy3_trainset(config, onet, traindl):
    # Get test results via feeding test_dataloader to net.
    onet.eval() # set the model to evaluation mode
    res_dict_list = [] #save all the results in dict
    with torch.no_grad():
        for X, img_ids in traindl: #
            batch_size = X.shape[0]
            X = X.to(config.device)

            Lt, St = onet.topu(X) # in this version softmax for topu is inside topu for two_class output.
            pred_labels = onet.predict_label(St)  # using the original input x's output to predict the label.
            if onet.fg_mark == 'down': #reverse the label if fg_mark is down.
                pred_labels = 1 - pred_labels

            for i in range(batch_size):  # iterate batch_size
                img_id = img_ids[i]
                tid = img_id.split('_')[-1]
                # if tid in bright_terrain_only_tid_list: #omit the bright terrain only images.
                #     continue
                if tid not in bright_terrain_only_tid_list:
                    continue
                res_dict = {}
                res_dict['img_id'] = img_ids[i]
                rgb   = X[i].permute(1, 2, 0).cpu().numpy()  # rgb[i] with 3 channels at last axis
                pred  = pred_labels[i].cpu().numpy()
                res_dict['coverage'] = np.sum(pred) / (pred.shape[0] * pred.shape[1])
                file_rgb = os.path.join(config.out_root, 'rgb_%s.png' % (img_ids[i]))
                img_rgb = Image.fromarray((rgb * 255).astype(np.uint8))
                img_rgb.save(file_rgb)
                file_pred = os.path.join(config.out_root, 'pred_%s.png' % (img_ids[i]))
                img_pred = Image.fromarray((pred * 255).astype(np.uint8))
                img_pred.save(file_pred)
                res_dict['pred_file'] = file_pred
                res_dict['rgb_file']  = file_rgb
                res_dict_list.append(res_dict)
    return res_dict_list

def write_image_to_excel(ws, img_file, cell, width, height):
    img = openpyxl.drawing.image.Image(img_file)
    img.anchor = cell
    img.width  = width
    img.height = height
    ws.add_image(img)

def save_results_to_excel(res_dict_list, excel_file):
    import pandas as pd
    #import xlsxwriter
    # save the results to excel file for further analysis.
    df = pd.DataFrame(res_dict_list)
    df = df.drop(columns=['rgb_file', 'pred_file'])
    df.to_excel(excel_file, index=False)

    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Sheet1']
    for i in range(len(res_dict_list)):
        img_name = res_dict_list[i]['rgb_file']
        ws['C1'] = 'rgb'
        write_image_to_excel(ws, img_name, 'C%d'%(i+2), 50, 50)

        ws['D1'] = 'pred'
        img_name = res_dict_list[i]['pred_file']
        write_image_to_excel(ws, img_name, 'D%d'%(i+2), 50, 50)

    wb.save(excel_file)

def divide_zy3_traindata(config):
    '''
    Divided the zy3 trainset into two parts: clouds_no_snow_ice_178 and bright_terrain_only_53, according to the Excel file.
    The Excel file is automatically generated by the function seg_on_zy3_trainset() and save_results_to_excel(),
    and then manually divided by the visual inspection.
    :param config:
    :return:
    '''
    import pandas as pd
    #excel_file = '/root/tip_onet_revision_gz16g/checkpoint/onet_vanilla_select_trainset_for_correct_clouds/zy3_trainset_divided_20240307.xlsx'
    excel_file = os.path.join(config.out_root, 'zy3_trainset_divided_20240307.xlsx')
    df = pd.read_excel(excel_file, sheet_name='clouds_no_snow_ice_178') # take the ids from the clouds_no_snow_ice_178 sheet
    img_no_snow_ids = df['img_id'].tolist()
    zy3_train_image_dict = cloud_model_zy3.prepare_cloud_traindata(config)
    zy3_train_clouds_no_snow_dict = {k: v for k, v in zy3_train_image_dict.items() if k in img_no_snow_ids}

    df = pd.read_excel(excel_file, sheet_name='bright_terrain_only_53') # take the ids from the bright_terrain_only_53 sheet
    img_bright_terrain_only_ids = df['img_id'].tolist()
    zy3_train_bright_terrain_only_dict = {k: v for k, v in zy3_train_image_dict.items() if k in img_bright_terrain_only_ids}

    print('get %d cloud images with no snow_ice in zy3_trainset, '% len(zy3_train_clouds_no_snow_dict))
    print('get %d images with no clouds in bright terrain only.  '% len(zy3_train_bright_terrain_only_dict))
    return zy3_train_clouds_no_snow_dict, zy3_train_bright_terrain_only_dict

def divide_zy3_testdata(config):
    '''
    Divided the zy3 trainset into three parts: normal_clouds, thin_clouds, and snow_clouds, according to the Excel file.
    The Excel file is automatically generated by the function seg_on_zy3_trainset() and save_results_to_excel(),
    and then manually divided by the visual inspection.
    :param config:
    :return:
    '''
    import pandas as pd
    #excel_file = '/root/tip_onet_revision_gz16g/checkpoint/onet_vanilla_select_trainset_for_correct_clouds/zy3_testset_divided_20240306.xlsx'
    excel_file = os.path.join(config.out_root, 'zy3_testset_divided_20240306.xlsx')
    df = pd.read_excel(excel_file, sheet_name='normal_clouds') # take the ids from sheet
    img_normal_ids = df['img_id'].tolist()
    zy3_test_image_dict = cloud_model_zy3.prepare_cloud_testdata(config)
    zy3_test_normal_clouds_dict = {k: v for k, v in zy3_test_image_dict.items() if k in img_normal_ids}

    df = pd.read_excel(excel_file, sheet_name='thin_cloud') # take the ids from  sheet
    img_thin_ids = df['img_id'].tolist()
    zy3_test_thin_cloud_dict = {k: v for k, v in zy3_test_image_dict.items() if k in img_thin_ids}

    df = pd.read_excel(excel_file, sheet_name='snow_cloud') # take the ids from  sheet
    img_snow_ids = df['img_id'].tolist()
    zy3_test_snow_cloud_dict = {k: v for k, v in zy3_test_image_dict.items() if k in img_snow_ids}

    print('get %d normal cloud images zy3_testset, '% len(zy3_test_normal_clouds_dict))
    print('get %d images with thin clouds,  '% len(zy3_test_thin_cloud_dict))
    print('get %d images with snow clouds.  '% len(zy3_test_snow_cloud_dict))
    return zy3_test_normal_clouds_dict, zy3_test_thin_cloud_dict, zy3_test_snow_cloud_dict


if __name__=='__main__':

    import pytz  # change timezone
    current_date = datetime.now(pytz.timezone('Asia/Shanghai'))  # change time zone to Beijing
    datehour_mark = '%04d%02d%02d_%02d' % (current_date.year, current_date.month, current_date.day, current_date.hour)
    print('Onet_pid: ',     os.getpid())
    print('current w_dir ', os.getcwd())

    config = conf_model.generate_config(
        './configs/select_trainset_for_correct_clouds_20240307.yml', dataset_name='zy3')
    print('checkpoint_directory:\n ', config.out_root)
    if not os.path.exists(config.out_root):
        os.makedirs(config.out_root)
    config.start_time = datetime.now(pytz.timezone('Asia/Shanghai'))  # change time zone to Beijing, for computing total training time.

    print(uti.config_to_str(config))  # print configuration to log file.
    config.log_file = config.log_file.split('.')[0] + '%s.log' % (datehour_mark)  # add time stamp to log file name
    logging.basicConfig(filename=config.log_file, encoding='utf-8', level=logging.INFO)

    zy3_train_clouds_no_snow_dict, zy3_train_bright_terrain_only_dict = divide_zy3_traindata(config)
    exit(0)

    # zy3_train_dataloader, zy3_test_dataloader = cloud_model_zy3.cloud_dataloader(config)
    zy3_train_image_dict = cloud_model_zy3.prepare_cloud_traindata(config)
    zy3_test_image_label_dict = cloud_model_zy3.prepare_cloud_testdata(config)

    zy3_aug_train_dataloader = cloud_model_zy3.cloud_dataloader_via_dict(config, zy3_train_image_dict, bsu=False, baug=True) # augmentation on images with no labels
    zy3_train_dataloader = cloud_model_zy3.cloud_dataloader_via_dict(config, zy3_train_image_dict, bsu=False, baug=False) # no augmentation on train set, but get labels.
    zy3_test_dataloader = cloud_model_zy3.cloud_dataloader_via_dict(config, zy3_test_image_label_dict, bsu=True,  baug=False) # no augmentation on test set, but get labels.


    # with output channel=2 for outc layer, should set n_classes=2 in onet_vanilla_outc.py
    # onet = onet_vanilla_outc.Onet(in_chns=3, n_classes=2, binit=True)  # in revised tip, onet_vanilla use RGB.
    # logging.info(uti.config_to_str(config))
    # logging.info(uti.count_parameters(onet, bverbose=False))  # count paramermters without detail names.
    # onet.to(config.device)

    # dict_info = torch.load(config.model_file, map_location=lambda storage, loc: storage)
    # onet.load_state_dict(dict_info["net"])
    # config.last_epoch = dict_info["save_epoch"]
    # config.save_epoch = config.last_epoch
    # uti_zy3_test.assign_fg_mark_v2(config, onet, zy3_test_dataloader)  # assign the fg_mark to onet.

    # res_dict_list = seg_on_zy3_trainset(config, onet, zy3_train_dataloader)
    # excel_file = os.path.join(config.out_root, 'seg_on_zy3_trainset_bright_terrain_only_%s.xlsx' % (datehour_mark))
    # save_results_to_excel(res_dict_list, excel_file)
    # print('Results saved to: ', excel_file)