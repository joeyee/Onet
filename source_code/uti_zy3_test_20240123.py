'''
utility functions for zy3 test
Created by ZhouYi@Linghai on 2024/01/23
based on the functions in 'selected_similar_set_for_onet_20240110.py'

'''

import os
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
import torch
import utils_20231218 as uti
import logging
torch.manual_seed(1981)
np.random.seed(1981)
torch.set_default_dtype(torch.float32)
import pytz # change timezone
from tqdm import tqdm
import pandas as pd
import openpyxl
from PIL import Image
import math

def print_sorted_results(res_dict_list):
    '''
    print the sorted results by acc or miou
    '''
    res_dict_list.sort(key=lambda x: x['acc'], reverse=True)
    for i in range(len(res_dict_list)):
        if 'dr' in res_dict_list[i]:
            print('%02d, img_id, %s, acc, %.4f, miou, %.4f, dr, %.4f, far, %.2E'
                  % (i, res_dict_list[i]['img_id'], res_dict_list[i]['acc'], res_dict_list[i]['miou'],
                     res_dict_list[i]['dr'], res_dict_list[i]['far']))
            logging.info('%02d, img_id, %s, acc, %.4f, miou, %.4f, dr, %.4f, far, %.2E'
                 % (i, res_dict_list[i]['img_id'], res_dict_list[i]['acc'], res_dict_list[i]['miou'],
                    res_dict_list[i]['dr'], res_dict_list[i]['far']))
        else:
            print('%02d, img_id, %s, acc, %.4f, miou, %.4f,' % (i, res_dict_list[i]['img_id'], res_dict_list[i]['acc'], res_dict_list[i]['miou']))
            logging.info('%02d, img_id, %s, acc, %.4f, miou, %.4f,' % (i, res_dict_list[i]['img_id'], res_dict_list[i]['acc'], res_dict_list[i]['miou']))
    return res_dict_list
def draw_test_res(config, test_loss, acc, miou, res_dict_list, draw_all=False):
    '''
    draw a batch in one figure or draw all the test_results in mutliple figures.
    '''
    #assert len(res_dict_list)%5==0 # figure contains 5 samples in 5 columns, in each column draw [rgb, label, pred_label, vt, vd]
    #assert (len(res_dict_list[0].keys()) == 8)  # because we have 8 keys in res_dict_list[0]

    num_samples = len(res_dict_list)
    indexs = np.arange(num_samples) # shuffle the indexs
    #np.random.shuffle(indexs) # shuffle the indexs
    if not draw_all:
        np.random.shuffle(indexs)  # shuffle the indexs
        indexs = indexs[:5] # draw a batch
        rounds = 1
    else:
        rounds = num_samples // 5
    keys = list(res_dict_list[0].keys())  # ['img_id', 'rgb', 'label', 'pred_label', 'vt', 'vd', 'acc', 'miou']

    for round in range(rounds): # each round draw five samples in one figure.
        fig, axs = plt.subplots(5, 5, gridspec_kw={'wspace': 0, 'hspace': 0}, figsize=(10, 10), sharex=True, sharey=True)
        for i in range(5):
            for j in range(5):
                if j==0:
                    img_id = res_dict_list[indexs[round*5+i]]['img_id']
                    tid = img_id.split('_')[-1]
                    sub_title = '%s\nacc:%.4f\nmiou:%.4f' % (img_id,
                                                             res_dict_list[indexs[round*5+i]]['acc'],
                                                             res_dict_list[indexs[round*5+i]]['miou'])
                    if 'dr' in res_dict_list[indexs[round*5+i]]:#get detection rate in res_dict_list
                        sub_title += '\ndr:%.4f\nfar:%.4f' % (res_dict_list[indexs[round*5+i]]['dr'], res_dict_list[indexs[round*5+i]]['far'])

                    axs[j, i].set_title(sub_title, fontsize=8)
                disp_array = res_dict_list[indexs[round*5+i]][keys[j+1]] # because the first key is img_id,  so j+1
                if np.all(disp_array==1) or np.all(disp_array==0):
                    disp_array[0,0] = 1 # avoid all 0 or all 1
                    disp_array[0,1] = 0
                axs[j, i].imshow(disp_array)
                axs[j, i].axis('off')
        fig.suptitle('zy3_rgb_gt_pred_vt_vd')
        if 'dr' in res_dict_list[0].keys():  # detection rate and far for simbg results
            mean_dr = np.mean([res_dict['dr'] for res_dict in res_dict_list])
            mean_far = np.mean([res_dict['far'] for res_dict in res_dict_list])
            if hasattr(config, 'last_epoch'):
                fig.savefig(os.path.join(config.out_root, '%s_epoch_%03d_round_%d_acc_%.4f_miou_%.4f_dr_%.4f_far_%.2E.png' %
                                            (config.model_name, config.last_epoch, round, acc, miou, mean_dr, mean_far)))
            else:
                fig.savefig(os.path.join(config.out_root, '%s_round_%d_acc_%.4f_miou_%.4f_dr_%.4f_far_%.2E.png' %
                                        (config.model_name, round, acc, miou, mean_dr, mean_far)))
        else:
            if hasattr(config, 'last_epoch'):
                fig.savefig(os.path.join(config.out_root, '%s_epoch_%03d_round_%d_loss_%.2E_acc_%.4f_miou_%.4f_dr_%.4f_far_%.2E.png' %
                                        (config.model_name, config.last_epoch, round, test_loss, acc, miou, mean_dr, mean_far)))
            else:
                fig.savefig(os.path.join(config.out_root, '%s_round_%d_loss_%.2E_acc_%.4f_miou_%.4f_dr_%.4f_far_%.2E.png' %
                                            (config.model_name, round, test_loss, acc, miou, mean_dr, mean_far)))
        plt.close(fig)

def assign_fg_mark(config, onet, test_dl):
    '''
    Given the test_dl(with gt_labels inside), assign the fg_mark to onet.
    :param onet:
    :param test_dl:
    :return:
    '''
    onet.eval()
    with torch.no_grad():
        bid = 0
        for X, labels, img_ids in test_dl:
            X = X.to(config.device)
            labels = labels.to(config.device)
            Lt, Vt, Ld, Vd, S = onet(X)
            pred_labels = onet.predict_label(S)
            if bid==0:
                onet.fg_mark = onet.determin_fg_mark(pred_labels, labels)
            else:
                assert onet.fg_mark == onet.determin_fg_mark(pred_labels, labels), 'fg_mark is not the same in different batches.'
            bid+=1
            break

def assign_fg_mark_v2(config, onet, test_dl):
    '''
    Given the test_dl(with gt_labels inside), assign the fg_mark to onet.
    :param onet:
    :param test_dl:
    :return:
    '''
    onet.eval()
    with torch.no_grad():
        bid = 0
        for X, labels, img_ids in test_dl:
            X = X.to(config.device)
            labels = labels.to(config.device)
            Lt, S = onet.topu(X)
            pred_labels = onet.predict_label(S)
            if bid==0:
                onet.fg_mark = onet.determin_fg_mark(pred_labels, labels)
                Y = uti.reorder_segmentation(pred_labels, labels)  # reorder the segmentation result.
                if torch.all(pred_labels == Y):
                    # since pred_labels = argmax(S, dim=1),
                    # St takes from S[:, 1, :, :], therefore the last channel is the fg_cloud channel.
                    onet.fg_mark = 'top'
                else:
                    assert (torch.all(pred_labels == 1 - Y))
                    onet.fg_mark = 'down'
            # else:
            #     assert onet.fg_mark == onet.determin_fg_mark(pred_labels, labels), 'fg_mark is not the same in different batches.'
            bid+=1
            break

def test_on_zy3_nail(config, onet, testdl_zy3_nail, draw_batch=False, draw_all=False):
    # Get test results via feeding test_dataloader to net.
    onet.eval()
    acc_batch_list  = []
    miou_batch_list = []
    dr_batch_list = [] # detection rate or probability of detection (pd) list
    far_batch_list = []
    res_dict_list = [] #save all the results in dict
    with torch.no_grad():
        test_loss_batch = []
        for X, label, img_ids in testdl_zy3_nail: #
            # if 'zy3_test_1706158599' in img_ids: # skip the image with no cloud
            #     print(img_ids)
            batch_size = X.shape[0]
            X = X.to(config.device)
            label = label.to(config.device)
            Lt, Vt, Ld, Vd, S = onet(X) #send rgb to onet for convnext backbone

            # pred_t and pred_d are used to display the raw prediction before softmax in show_unet_adversarial_v2().
            pred_t = Vt.squeeze(dim=1)
            pred_d = Vd.squeeze(dim=1)
            #pred_labels = torch.argmax(S, dim=1)
            pred_labels = onet.predict_label(S)
            Y = uti.reorder_segmentation(pred_labels, label)
            St = S[:, 0, :, :].unsqueeze(dim=1)
            Sd = S[:, 1, :, :].unsqueeze(dim=1)

            loss = onet.compute_loss(Lt,St,Ld, Sd)

            test_loss_batch.append(loss.item())
            batch_acc, batch_miou = uti.evaluate_segmentation(Y, label, gt_k=2)
            # acc_batch_list.append(batch_acc)
            # miou_batch_list.append(batch_miou)
            for i in range(batch_size):  # iterate batch_size
                res_dict = {}
                res_dict['img_id'] = img_ids[i]
                # if img_ids[i] == 'zy3_test_1706158599':
                #     print(img_ids[i])
                res_dict['rgb'] = X[i].permute(1, 2, 0).cpu().numpy()  # rgb[i] with 3 channels at last axis
                res_dict['label'] = label[i].cpu().numpy()
                #Y = uti.reorder_segmentation(pred_labels[i], label[i])
                res_dict['pred_label'] = Y[i].cpu().numpy()
                res_dict['vt'] = pred_t[i].cpu().numpy()
                res_dict['vd'] = pred_d[i].cpu().numpy()
                #res_dict['acc'], res_dict['miou'] = uti.evaluate_segmentation(Y[i], label[i], gt_k=2)
                res_dict['acc'], res_dict['miou'], res_dict['dr'], res_dict['far'], _ = uti.evaluate_nau_segmentation_v2(Y[i], label[i], gt_k=2)
                # if math.isnan(res_dict['miou']):
                #     print('miou is nan')
                acc_batch_list.append(res_dict['acc'])
                dr_batch_list.append(res_dict['dr'])
                far_batch_list.append(res_dict['far'])
                miou_batch_list.append(res_dict['miou'])
                res_dict_list.append(res_dict)
        test_loss_epoch = np.mean(test_loss_batch)
        acc_epoch = np.array(acc_batch_list).mean()
        miou_epoch = np.array(miou_batch_list).mean()
        dr_epoch = np.array(dr_batch_list).mean()
        far_epoch = np.array(far_batch_list).mean()

    if draw_batch or draw_all: # when draw samples output results.
        res_dict_list = print_sorted_results(res_dict_list)  # print the sorted results of each file by acc or miou
        print('---Test results on ZY3_thumbnails_cloud_segmentation:-------')
        if hasattr(config, 'save_epoch'):
            print('Best model saved at ', config.save_epoch)
        print('Test JSD loss: %.4f' % test_loss_epoch)
        print('Accuracy %.4f, mIoU %.4f, dr %.4f, far %.4f'% (acc_epoch, miou_epoch, dr_epoch, far_epoch))
        print('Time cost: ', datetime.now(pytz.timezone('Asia/Shanghai'))-config.start_time)
        print('-----------------------------------------------------------')
        logging.info('---Test results on ZY3_thumbnails_cloud_segmentation:-------')
        if hasattr(config, 'save_epoch'):
            logging.info('Best model saved at %d' % config.save_epoch)
        logging.info('Accuracy %.4f, mIoU %.4f, dr %.4f, far %.4f' % (acc_epoch, miou_epoch, dr_epoch, far_epoch))
        logging.info('Time cost: %s' % str(datetime.now(pytz.timezone('Asia/Shanghai')) - config.start_time))
        logging.info('-----------------------------------------------------------')
        # normally draw a random batch. if draw_all is True, draw all the results.
        draw_test_res(config, test_loss_epoch, acc_epoch, miou_epoch, res_dict_list, draw_all)

    #select_test_set(res_dict_list) # save the lower acc images to selected_dict_list.pt
    #return test_loss_epoch, acc_epoch, miou_epoch
    return test_loss_epoch, acc_epoch, miou_epoch, dr_epoch, far_epoch

def test_on_zy3_nail_v2(config, onet, testdl_zy3_nail, draw_batch=False, draw_all=False):
    # Get test results via feeding test_dataloader to net.
    onet.eval()
    acc_batch_list  = []
    miou_batch_list = []
    res_dict_list = [] #save all the results in dict
    with torch.no_grad():
        test_loss_batch = []
        for X, label, img_ids in testdl_zy3_nail: #
            # if 'zy3_test_1706158599' in img_ids: # skip the image with no cloud
            #     print(img_ids)
            batch_size = X.shape[0]
            X = X.to(config.device)
            label = label.to(config.device)

            Lt, St = onet.topu(X) # in this version softmax for topu is inside topu for two_class output.
            Vt = St[:, 1, :, :].unsqueeze(1)
            Ld, Sd = onet.topu(1-X)  # negative_augmentation for unsupervised learning.
            Vd = Sd[:, 1, :, :].unsqueeze(1)
            S = onet.softmax(torch.cat([Vt, Vd], dim=1)) #final score map for two classes.
            Sf = S[:, 0, :, :].unsqueeze(dim=1)
            Sb = S[:, 1, :, :].unsqueeze(dim=1)
            loss = onet.compute_loss(Lt, Sf, Ld, Sb)

            # pred_t and pred_d are used to display the raw prediction before softmax in show_unet_adversarial_v2().
            pred_t = Vt.squeeze(dim=1)
            pred_d = Vd.squeeze(dim=1)
            pred_labels = torch.argmax(S, dim=1)
            #pred_labels = onet.predict_label(St)  # using the original input x's output to predict the label.
            Y = uti.reorder_segmentation(pred_labels, label)


            test_loss_batch.append(loss.item())
            batch_acc, batch_miou = uti.evaluate_segmentation(Y, label, gt_k=2)
            # acc_batch_list.append(batch_acc)
            # miou_batch_list.append(batch_miou)
            for i in range(batch_size):  # iterate batch_size
                res_dict = {}
                res_dict['img_id'] = img_ids[i]
                res_dict['rgb'] = X[i].permute(1, 2, 0).cpu().numpy()  # rgb[i] with 3 channels at last axis
                res_dict['label'] = label[i].cpu().numpy()
                #Y = uti.reorder_segmentation(pred_labels[i], label[i])
                res_dict['pred_label'] = Y[i].cpu().numpy()
                res_dict['vt'] = pred_t[i].cpu().numpy()
                res_dict['vd'] = pred_d[i].cpu().numpy()
                res_dict['acc'], res_dict['miou'] = uti.evaluate_segmentation(Y[i], label[i], gt_k=2)
                acc_batch_list.append(res_dict['acc'])
                miou_batch_list.append(res_dict['miou'])
                res_dict_list.append(res_dict)
        test_loss_epoch = np.mean(test_loss_batch)
        acc_epoch = np.array(acc_batch_list).mean()
        miou_epoch = np.array(miou_batch_list).mean()

    if draw_batch or draw_all: # when draw samples output results.
        res_dict_list = print_sorted_results(res_dict_list)  # print the sorted results of each file by acc or miou
        print('---Test results on ZY3_thumbnails_cloud_segmentation:-------')
        print('Best model saved at ', config.save_epoch)
        print('Test JSD loss: %.4f' % test_loss_epoch)
        print('Accuracy %.4f, mIoU %.4f'% (acc_epoch, miou_epoch))
        print('Time cost: ', datetime.now(pytz.timezone('Asia/Shanghai'))-config.start_time)
        print('-----------------------------------------------------------')
        logging.info('---Test results on ZY3_thumbnails_cloud_segmentation:-------')
        logging.info('Best model saved at %d' % config.save_epoch)
        logging.info('Accuracy %.4f, mIoU %.4f' % (acc_epoch, miou_epoch))
        logging.info('Time cost: %s' % str(datetime.now(pytz.timezone('Asia/Shanghai')) - config.start_time))
        logging.info('-----------------------------------------------------------')
        # normally draw a random batch. if draw_all is True, draw all the results.
        draw_test_res(config, test_loss_epoch, acc_epoch, miou_epoch, res_dict_list, draw_all)

    #select_test_set(res_dict_list) # save the lower acc images to selected_dict_list.pt
    return test_loss_epoch, acc_epoch, miou_epoch

def get_divided_test_list():
    '''get the divided test list from the Excel file and return them together as a dict.'''
    #excel_file = '/root/tip_onet_revision_gz16g/checkpoint/onet_vanilla_select_trainset_for_correct_clouds/zy3_testset_divided_20240306.xlsx'
    excel_file = '/root/onet_github/checkpoint/zy3/onet_vanilla/zy3_testset_divided_20240306.xlsx'
    print('load the divided information from %s'%excel_file)
    df = pd.read_excel(excel_file, sheet_name='normal_clouds') # take the ids from sheet
    img_normal_ids = df['img_id'].tolist()

    df = pd.read_excel(excel_file, sheet_name='thin_cloud') # take the ids from  sheet
    img_thin_ids = df['img_id'].tolist()

    df = pd.read_excel(excel_file, sheet_name='snow_cloud') # take the ids from  sheet
    img_snow_ids = df['img_id'].tolist()

    return {'normal_cloud': img_normal_ids, 'thin_cloud': img_thin_ids, 'snow_cloud': img_snow_ids}

def save_zy3_test_results_to_excel(config, onet, testdl_zy3, divided_list_dict, bsave_excel=False):
    # Save the res_dict_list to an excel file in divided format.
    # 1. all the results in res_dict_list are saved in an excel file's 'all' sheet.
    # 2. divide the samples into three groupts: 0 for normal cloud without snow, 1 for thin clouds, 2 for clouds with snow.
    onet.eval()
    acc_list = []
    miou_list = []
    loss_list = []
    res_dict_list = []  # save all the results in dict
    with torch.no_grad():
        for X, label, img_ids in testdl_zy3:  #
            batch_size = X.shape[0]
            X = X.to(config.device)
            label = label.to(config.device)

            Lt, Vt, Ld, Vd, S = onet(X)  # send rgb to onet for convnext backbone

            pred_t = Vt.squeeze(dim=1)
            pred_d = Vd.squeeze(dim=1)
            # pred_labels = torch.argmax(S, dim=1)
            pred_labels = onet.predict_label(S)
            Y = uti.reorder_segmentation(pred_labels, label)
            St = S[:, 0, :, :].unsqueeze(dim=1)
            Sd = S[:, 1, :, :].unsqueeze(dim=1)

            loss = onet.compute_loss(Lt, St, Ld, Sd)

            # pred_labels = onet.predict_label(St)  # using the original input x's output to predict the label.
            Y = uti.reorder_segmentation(pred_labels, label)
            loss_list.append(loss.item())
            for i in range(batch_size):  # iterate batch_size
                res_dict = {}
                img_id = img_ids[i]
                res_dict['img_id'] = img_id
                res_dict['rgb'] = X[i].permute(1, 2, 0).cpu().numpy()  # rgb[i] with 3 channels at last axis
                res_dict['label'] = label[i].cpu().numpy()
                # Y = uti.reorder_segmentation(pred_labels[i], label[i])
                res_dict['pred_label'] = Y[i].cpu().numpy()
                res_dict['vt'] = pred_t[i].cpu().numpy()
                res_dict['vd'] = pred_d[i].cpu().numpy()
                res_dict['acc'], res_dict['miou'] = uti.evaluate_segmentation(Y[i], label[i], gt_k=2)
                acc_list.append(res_dict['acc'])
                miou_list.append(res_dict['miou'])

                if img_id in divided_list_dict['normal_cloud']:
                    res_dict['group'] = 0
                elif img_id in divided_list_dict['thin_cloud']:
                    res_dict['group'] = 1
                elif img_id in divided_list_dict['snow_cloud']:
                    res_dict['group'] = 2
                else:
                    print('img_id not in any divided group: ', img_id)
                if bsave_excel:
                    file_rgb = os.path.join(config.out_root, '%s_rgb.png' % (img_ids[i]))
                    if not os.path.exists(file_rgb):
                        img_rgb = Image.fromarray((res_dict['rgb'] * 255).astype(np.uint8))
                        img_rgb.save(file_rgb)
                    file_label = os.path.join(config.out_root, '%s_label.png' % (img_ids[i]))
                    if not os.path.exists(file_label):
                        img_label = Image.fromarray((res_dict['label'] * 255).astype(np.uint8))
                        img_label.save(file_label)
                    file_pred = os.path.join(config.out_root, '%s_pred.png' % (img_ids[i]))
                    img_pred = Image.fromarray((res_dict['pred_label'] * 255).astype(np.uint8))
                    img_pred.save(file_pred)
                    file_vt = os.path.join(config.out_root, '%s_vt.png' % (img_ids[i]))
                    img_vt = Image.fromarray((res_dict['vt'] * 255).astype(np.uint8))
                    img_vt.save(file_vt)
                    file_vd = os.path.join(config.out_root, '%s_vd.png' % (img_ids[i]))
                    img_vd = Image.fromarray((res_dict['vd'] * 255).astype(np.uint8))
                    img_vd.save(file_vd)

                    res_dict['label_file'] = file_label
                    res_dict['pred_file'] = file_pred
                    res_dict['rgb_file'] = file_rgb
                    res_dict['vt_file'] = file_vt
                    res_dict['vd_file'] = file_vd
                res_dict_list.append(res_dict)
    test_loss = np.mean(loss_list)
    acc = np.array(acc_list).mean()
    miou = np.array(miou_list).mean()
    print('Overall testset Accuracy %.4f, mIoU %.4f' % (acc, miou))
    logging.info('Overall testset Accuracy %.4f, mIoU %.4f' % (acc, miou))
    # compute the mean acc and miou for each group.
    mean_acc = [0, 0, 0]
    mean_miou = [0, 0, 0]
    num_samples = [0, 0, 0]
    acc_divided_dict = {}
    miou_divided_dict = {}
    mark = ['normal_cloud', 'thin_cloud', 'snow_cloud']
    for res_dict in res_dict_list:
        group = res_dict['group']
        mean_acc[group] += res_dict['acc']
        mean_miou[group] += res_dict['miou']
        num_samples[group] += 1
    for i in range(3):
        mean_acc[i] /= num_samples[i]
        mean_miou[i] /= num_samples[i]
        acc_divided_dict[mark[i]] = mean_acc[i]
        miou_divided_dict[mark[i]] = mean_miou[i]
        print('%s mean_acc  : %.4f, mean_miou : %.4f, num_samples: %d' % (
        mark[i], mean_acc[i], mean_miou[i], num_samples[i]))
        logging.info('%s mean_acc  : %.4f, mean_miou : %.4f, num_samples: %d' % (
        mark[i], mean_acc[i], mean_miou[i], num_samples[i]))
    if bsave_excel:
        res_excel_file = os.path.join(config.out_root, config.res_excel_file)
        time_mark = datetime.now().strftime('%Y%m%d%H')
        res_excel_file = res_excel_file.replace('.xlsx', '_%s.xlsx' % time_mark)
        save_results_to_excel(res_dict_list, res_excel_file)
    # return test_loss, acc, miou, acc_divided_dict, miou_divided_dict
    return res_dict_list

def test_on_zy3_nail_v3(config, onet, testdl_zy3, divided_list_dict, bsave_excel=False):
    # Save the res_dict_list to an excel file in divided format.
    # 1. all the results in res_dict_list are saved in an excel file's 'all' sheet.
    # 2. divide the samples into three groupts: 0 for normal cloud without snow, 1 for thin clouds, 2 for clouds with snow.
    onet.eval()
    acc_list  = []
    miou_list = []
    loss_list= []
    res_dict_list = [] #save all the results in dict
    with torch.no_grad():
        for X, label, img_ids in testdl_zy3: #
            batch_size = X.shape[0]
            X = X.to(config.device)
            label = label.to(config.device)

            Lt, St = onet.topu(X) # in this version softmax for topu is inside topu for two_class output.
            Vt = St[:, 1, :, :].unsqueeze(1)
            Ld, Sd = onet.topu(1-X)  # negative_augmentation for unsupervised learning.
            Vd = Sd[:, 1, :, :].unsqueeze(1)
            S = onet.softmax(torch.cat([Vt, Vd], dim=1)) #final score map for two classes.
            Sf = S[:, 0, :, :].unsqueeze(dim=1)
            Sb = S[:, 1, :, :].unsqueeze(dim=1)
            loss = onet.compute_loss(Lt, Sf, Ld, Sb)

            # pred_t and pred_d are used to display the raw prediction before softmax in show_unet_adversarial_v2().
            pred_t = Vt.squeeze(dim=1)
            pred_d = Vd.squeeze(dim=1)
            pred_labels = torch.argmax(S, dim=1)
            #pred_labels = onet.predict_label(St)  # using the original input x's output to predict the label.
            Y = uti.reorder_segmentation(pred_labels, label)
            loss_list.append(loss.item())
            for i in range(batch_size):  # iterate batch_size
                res_dict = {}
                img_id = img_ids[i]
                res_dict['img_id'] = img_id
                res_dict['rgb'] = X[i].permute(1, 2, 0).cpu().numpy()  # rgb[i] with 3 channels at last axis
                res_dict['label'] = label[i].cpu().numpy()
                #Y = uti.reorder_segmentation(pred_labels[i], label[i])
                res_dict['pred_label'] = Y[i].cpu().numpy()
                res_dict['vt'] = pred_t[i].cpu().numpy()
                res_dict['vd'] = pred_d[i].cpu().numpy()
                res_dict['acc'], res_dict['miou'] = uti.evaluate_segmentation(Y[i], label[i], gt_k=2)
                acc_list.append(res_dict['acc'])
                miou_list.append(res_dict['miou'])

                if img_id in divided_list_dict['normal_cloud']:
                    res_dict['group'] = 0
                elif img_id in divided_list_dict['thin_cloud']:
                    res_dict['group'] = 1
                elif img_id in divided_list_dict['snow_cloud']:
                    res_dict['group'] = 2
                else:
                    print('img_id not in any divided group: ', img_id)
                if bsave_excel:
                    file_rgb = os.path.join(config.out_root, '%s_rgb.png' % (img_ids[i]))
                    if not os.path.exists(file_rgb):
                        img_rgb = Image.fromarray((res_dict['rgb'] * 255).astype(np.uint8))
                        img_rgb.save(file_rgb)
                    file_label = os.path.join(config.out_root, '%s_label.png' % (img_ids[i]))
                    if not os.path.exists(file_label):
                        img_label = Image.fromarray((res_dict['label']  * 255).astype(np.uint8))
                        img_label.save(file_label)
                    file_pred = os.path.join(config.out_root, '%s_pred.png' % (img_ids[i]))
                    img_pred = Image.fromarray((res_dict['pred_label'] * 255).astype(np.uint8))
                    img_pred.save(file_pred)
                    file_vt = os.path.join(config.out_root, '%s_vt.png' % (img_ids[i]))
                    img_vt = Image.fromarray((res_dict['vt'] * 255).astype(np.uint8))
                    img_vt.save(file_vt)
                    file_vd = os.path.join(config.out_root, '%s_vd.png' % (img_ids[i]))
                    img_vd = Image.fromarray((res_dict['vd'] * 255).astype(np.uint8))
                    img_vd.save(file_vd)

                    res_dict['label_file'] = file_label
                    res_dict['pred_file'] = file_pred
                    res_dict['rgb_file'] = file_rgb
                    res_dict['vt_file'] = file_vt
                    res_dict['vd_file'] = file_vd
                res_dict_list.append(res_dict)
    test_loss = np.mean(loss_list)
    acc = np.array(acc_list).mean()
    miou = np.array(miou_list).mean()
    print('Overall testset Accuracy %.4f, mIoU %.4f' % (acc, miou))
    logging.info('Overall testset Accuracy %.4f, mIoU %.4f' % (acc, miou))
    # compute the mean acc and miou for each group.
    mean_acc = [0, 0, 0]
    mean_miou = [0, 0, 0]
    num_samples = [0, 0, 0]
    acc_divided_dict={}
    miou_divided_dict={}
    mark = ['normal_cloud', 'thin_cloud', 'snow_cloud']
    for res_dict in res_dict_list:
        group = res_dict['group']
        mean_acc[group] += res_dict['acc']
        mean_miou[group] += res_dict['miou']
        num_samples[group] += 1
    for i in range(3):
        mean_acc[i] /= num_samples[i]
        mean_miou[i] /= num_samples[i]
        acc_divided_dict[mark[i]] = mean_acc[i]
        miou_divided_dict[mark[i]] = mean_miou[i]
        print('%s mean_acc  : %.4f, mean_miou : %.4f, num_samples: %d' % (mark[i], mean_acc[i], mean_miou[i], num_samples[i]))
        logging.info('%s mean_acc  : %.4f, mean_miou : %.4f, num_samples: %d' % (mark[i], mean_acc[i], mean_miou[i], num_samples[i]))
    if bsave_excel:
        res_excel_file = os.path.join(config.out_root, config.res_excel_file)
        time_mark = datetime.now().strftime('%Y%m%d%H')
        res_excel_file = res_excel_file.replace('.xlsx', '_%s.xlsx'%time_mark)
        save_results_to_excel(res_dict_list, res_excel_file)
    #return test_loss, acc, miou, acc_divided_dict, miou_divided_dict
    return  res_dict_list

def save_image_to_cell(ws, img_file, cell):
    from openpyxl.drawing.image import Image as xlImage
    import numpy as np
    from PIL import Image
    import os

    img = openpyxl.drawing.image.Image(img_file)
    img.width =50
    img.height=50
    img.anchor=cell
    # Add the image to the worksheet
    ws.add_image(img)


def save_results_to_excel(res_dict_list, excel_file):
    import pandas as pd
    from openpyxl.drawing.image import Image as xlImage
    from io import BytesIO
    from PIL import Image
    #import xlsxwriter
    # save the results to excel file for further analysis.
    if os.path.exists(excel_file):
        time_mark = datetime.now().strftime('%Y%m%d%H%M%S')
        excel_file = excel_file.replace('.xlsx', '_%s.xlsx'%time_mark)
    df = pd.DataFrame(res_dict_list)
    df = df.drop(columns=['vt', 'vd', 'rgb', 'label', 'pred_label', 'rgb_file', 'label_file', 'pred_file'])
    df.to_excel(excel_file, index=False)

    #saving images to excel file, replace the arrays with image previews.
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Sheet1']
    for i in range(len(res_dict_list)):
        ws['I1'] = 'rgb'
        save_image_to_cell(ws, res_dict_list[i]['rgb_file'], f'I{i+2}')
        ws['J1'] = 'label'
        save_image_to_cell(ws, res_dict_list[i]['label_file'], f'J{i+2}')
        ws['K1'] = 'pred'
        save_image_to_cell(ws, res_dict_list[i]['pred_file'], f'K{i+2}')
        ws['L1'] = 'vt'
        save_image_to_cell(ws, res_dict_list[i]['vt_file'], f'L{i+2}')
        ws['M1'] = 'vd'
        save_image_to_cell(ws, res_dict_list[i]['vd_file'], f'M{i+2}')
    wb.save(excel_file)
    # remove the temp files
    for i in range(len(res_dict_list)):
        os.remove(res_dict_list[i]['rgb_file'])
        os.remove(res_dict_list[i]['label_file'])
        os.remove(res_dict_list[i]['pred_file'])
        os.remove(res_dict_list[i]['vt_file'])
        os.remove(res_dict_list[i]['vd_file'])
    print('save the results in res_dict_list to %s' % excel_file)


def test_on_rayleigh_bg(config, onet, testdl_rayleigh_bg, draw_batch=False, draw_all=False):
    # Get test results via feeding test_dataloader to net.
    onet.eval()
    acc_batch_list  = []
    miou_batch_list = []
    dr_batch_list = []
    far_batch_list= []
    res_dict_list = [] #save all the results in dict
    with torch.no_grad():
        test_loss_batch = []
        for X, labels, psnrs in testdl_rayleigh_bg: #
            # if 'zy3_test_1706158599' in img_ids: # skip the image with no cloud
            #     print(img_ids)
            batch_size = X.shape[0]
            X = X.to(config.device)
            labels = labels.to(config.device)

            Lt, Vt, St = onet(X)
            Ld, Vd, Sd = onet(1-X)  # negative_augmentation for unsupervised learning.
            if config.gt_k == 1: # onet1.0's prediction
                V = torch.concat([Vt, Vd], dim=1)
                S = onet.softmax(V)  # gradient  backpropagation are needed, so not use onet.get_label()
                St = S[:, 0, :, :].unsqueeze(dim=1)
                Sd = S[:, 1, :, :].unsqueeze(dim=1)
                loss = onet.compute_loss(Lt, St, Ld, Sd)
                pred_t = Vt.squeeze(dim=1)
                pred_d = Vd.squeeze(dim=1)
                #pred_labels = torch.argmax(S, dim=1)
                pred_labels = onet.predict_label(S)  # using the original input x's output to predict the label.
            else: # onet2.0's prediction
                assert (config.gt_k == 2)  # two class output
                loss = onet.compute_loss_v2(Lt, St, Ld, Sd)
                # pred_t and pred_d are used to display the raw prediction before softmax in show_unet_adversarial_v2().
                pred_t = Vt[:,1,:,:]
                pred_d = Vd[:,1,:,:]
                #pred_labels = torch.argmax(S, dim=1)
                pred_labels = onet.predict_label(St)  # using the original input x's output to predict the label.

            Y = uti.reorder_segmentation(pred_labels, labels)
            test_loss_batch.append(loss.item())
            if (not draw_batch) and (not draw_all): #normal test without drawing results
                #batch_acc, batch_miou = uti.evaluate_segmentation(Y, labels, gt_k=2)
                batch_acc, batch_miou, batch_dr, batch_far = uti.evaluate_nau_segmentation_v2(Y, labels, gt_k=2)
                acc_batch_list.append(batch_acc)
                miou_batch_list.append(batch_miou)
                dr_batch_list.append(batch_dr)
                far_batch_list.append(batch_far)
            else: # test with drawing requirement
                for i in range(batch_size):  # iterate batch_size
                    res_dict = {}
                    res_dict['img_id'] = 'psnr_%d'%psnrs[i]
                    res_dict['rgb']    = X[i].permute(1, 2, 0).cpu().numpy()  # rgb[i] with 3 channels at last axis
                    res_dict['label']  = labels[i].cpu().numpy()
                    res_dict['pred_label'] = Y[i].cpu().numpy()
                    res_dict['vt'] = pred_t[i].cpu().numpy()
                    res_dict['vd'] = pred_d[i].cpu().numpy()
                    res_dict['acc'], res_dict['miou'], res_dict['dr'], res_dict['far'] = uti.evaluate_nau_segmentation_v2(Y[i], labels[i], gt_k=2)
                    acc_batch_list.append(res_dict['acc'])
                    miou_batch_list.append(res_dict['miou'])
                    dr_batch_list.append(res_dict['dr'])
                    far_batch_list.append(res_dict['far'])
                    res_dict_list.append(res_dict)
        test_loss_epoch = np.mean(test_loss_batch)
        acc_epoch = np.array(acc_batch_list).mean()
        miou_epoch = np.array(miou_batch_list).mean()
        dr_epoch = np.array(dr_batch_list).mean()
        far_epoch= np.array(far_batch_list).mean()

    if draw_batch or draw_all: # when draw samples output results.
        res_dict_list = print_sorted_results(res_dict_list)  # print the sorted results of each file by acc or miou
        # print('---Test results on Rayleigh_clutter_segmentation:-------')
        # print('Best model saved at ', config.save_epoch)
        # print('Test JSD loss: %.4f' % test_loss_epoch)
        # print('Accuracy %.4f, mIoU %.4f, dr %.4f, far %.4f'% (acc_epoch, miou_epoch, dr_epoch, far_epoch))
        # print('Time cost: ', datetime.now(pytz.timezone('Asia/Shanghai'))-config.start_time)
        # print('-----------------------------------------------------------')
        logging.info('---Test results on Rayleigh_clutter_segmentation:-------')
        logging.info('Best model saved at %d' % config.save_epoch)
        logging.info('Accuracy %.4f, mIoU %.4f, dr %.4f, far %.4f'% (acc_epoch, miou_epoch, dr_epoch, far_epoch))
        logging.info('Time cost: %s' % str(datetime.now(pytz.timezone('Asia/Shanghai')) - config.start_time))
        logging.info('-----------------------------------------------------------')
        # normally draw a random batch. if draw_all is True, draw all the results.
        draw_test_res(config, test_loss_epoch, acc_epoch, miou_epoch, res_dict_list, draw_all)

    #select_test_set(res_dict_list) # save the lower acc images to selected_dict_list.pt
    return test_loss_epoch, acc_epoch, miou_epoch, dr_epoch, far_epoch

def get_res_from_log(logfile):
    '''get the test results from the log file'''
    # Read the log file
    with open(logfile, 'r') as file:
        lines = file.readlines()
    # Reverse the order of the lines
    lines = reversed(lines)
    acc, miou, dr, far = (-1, -1, -1, -1) #not exits
    # Find the line with the relevant information
    for line in lines:
        if 'Accuracy' in line:
            # Split the line into individual values
            values = line.strip().split(', ')
            # Extract the values in reverse order
            if 'far' in line:
                far = float(values[3].split(' ')[1])
                dr = float(values[2].split(' ')[1])
            miou = float(values[1].split(' ')[1])
            acc = float(values[0].split(' ')[1])
            # Exit the loop once the values are found
            break

    if dr==-1:
        print('Attention, no results in log file!!!')

    return acc, miou, dr, far

import re
def get_dr_far_list_from_log(logfile):
    '''get the test results from the log file'''
    # Read the log file
    #print('gets epoch dr and far in %s'%logfile)
    with open(logfile, 'r') as file:
        lines = file.readlines()
    # Reverse the order of the lines
    epoch_list = []
    dr_list = []
    far_list = []
    # Find the line with the relevant information
    for line in lines:
        pattern = r'===Epoch:\s*(\d+).+dr\s+(\d+\.\d+),\s*far\s+(\d+\.\d+E[-+]?\d+)'
        match = re.search(pattern, line, re.IGNORECASE)
        if match:
            epoch = int(match.group(1))
            dr = float(match.group(2))
            far = float(match.group(3))
            epoch_list.append(epoch)
            dr_list.append(dr)
            far_list.append(far)
    nrange = torch.arange(0,301)
    assert(torch.all(torch.tensor(epoch_list)==nrange)), 'epoch_list is not continuous from 0 to 300.'
    return dr_list, far_list

def compute_ave_pd_in_order_of_magnitude(dr_list, far_list):
    '''compute the average pd in order of far's magnitude'''
    drs = torch.tensor(dr_list)
    fars = torch.tensor(far_list)
    assert fars.max() < 1 and fars.min()>=0, 'far should be less than 1.'
    far_order_magnitude =  (torch.log10(fars)).to(torch.int) - 1
    min_order = far_order_magnitude.min()
    max_order = 0#far_order_magnitude.max()
    pd_far_dict  ={}
    for order in range(min_order, max_order): #from min to -1.
        idx = (far_order_magnitude==order)
        if torch.all(idx==False): #omit the null order
            continue
        ave_pd = drs[idx].mean().item()
        pd_far_dict[order] = ave_pd
    return pd_far_dict
def format_latex_table(data):
    """
        Formats a list of dictionaries into an aligned LaTeX table.
        data = list[dict], where dict = {'key1': value1, 'key2': value2, ...}
        """
    # Get the keys from the first dictionary
    keys = list(data[0].keys())

    # Determine the maximum length of each key
    max_key_lengths = {key: len(key) for key in keys}
    for item in data:
        for key, value in item.items():
            max_key_lengths[key] = max(max_key_lengths[key], len(str(value)))

    # Create the table header
    table_header = " & ".join(f"{key.ljust(max_key_lengths[key])}" for key in keys) + " \\\\"
    table_header += "\n\\hline"

    # Create the table rows
    table_rows = []
    for item in data:
        row_values = [f"{str(item[key]).rjust(max_key_lengths[key])}" for key in keys]
        row_text = " & ".join(row_values) + " \\\\"
        table_rows.append(row_text)

    # Combine the header and rows into a complete table
    table = "\\begin{tabular}{" + "l" * len(keys) + "}\n"
    table += table_header + "\n"
    table += "\n".join(table_rows)
    table += "\n\\end{tabular}"

    return table